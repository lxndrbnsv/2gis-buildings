from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def load_queries() -> list:
    with open("address.txt", "r") as text_file:
        text_data = text_file.readlines()

    return [t.replace("\n", "") for t in text_data]


def get_building_info(query: str) -> str:
    url = f"https://2gis.ru/search/{query}"

    browser.get(url)

    building_card = WebDriverWait(browser, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "_1kf6gff"))
    )

    building_type = building_card.find_element(By.CLASS_NAME, "_1idnaau")

    return building_type.text


options = Options()
service = Service()
browser = webdriver.Chrome(service=service, options=options)
action = ActionChains(browser)
browser.set_window_size("1000", "800")


def write_data(row: tuple) -> None:
    wb = load_workbook(filename="results.xlsx")
    ws = wb[wb.sheetnames[0]]

    print(ws.min_row)
    print(ws.max_row)

    line_to_write = ws.max_row + 1
    ws[f"A{line_to_write}"] = row[0]
    ws[f"B{line_to_write}"] = row[1]
    wb.save(filename="results.xlsx")


def create_workbook() -> None:
    wb = Workbook()
    ws = wb[wb.sheetnames[0]]

    ws["A1"] = "Адрес"
    ws["B1"] = "Инфо"

    wb.save(filename="results.xlsx")


def main():
    create_workbook()
    queries = load_queries()
    for num, query in enumerate(queries):
        print(f"{num + 1} из {len(queries)}: {query}")
        building_info = get_building_info(query)
        write_data((query, building_info))


if __name__ == "__main__":
    main()
